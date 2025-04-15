import shutil
import pdb
import os
import argparse
from datetime   import datetime, date, timedelta, time
from copy       import copy

from openpyxl                import load_workbook
from openpyxl.styles         import Font, Border, Side,  Alignment
from openpyxl.cell.text      import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock, TextBlock
from openpyxl.utils import column_index_from_string, get_column_letter

from meituan.main       import get_meituanSum, mt_status
from douyin.main        import get_douyinSum
from operation.main     import resolve_operation_data
from operation import ThirdParty
from operation import elecdata as electron
from specialFee      import main as specialFee
from tools              import env
from tools.logger import get_logger

logger = get_logger(__name__)
logger.info('程序运行中')

parser = argparse.ArgumentParser(description="日报表自动化套件")
# 必须参数
#parser.add_argument("input", help="输入文件路径")
#parser.add_argument("output", help="输出文件路径")
# 可选参数
parser.add_argument("-v", "--verbose", action="store_true", help="显示详细日志")
parser.add_argument("-dc", "--dateconfig", action="store_true", help="使用配置文件的日期")
parser.add_argument("-now", "--today", action="store_true", help="使用今天作为日期")
parser.add_argument("-ne", "--no-elec", action="store_true", help="不输入电表数据")
parser.add_argument("-dt", "--date", type=str  , help="指定日期")

args = parser.parse_args()

yesterday = date.today() - timedelta(days=1)
working_datetime_date = yesterday
working_date_str      = yesterday.strftime('%Y-%m-%d')
working_datetime      = datetime.combine(yesterday, datetime.min.time())



if args.dateconfig:
    working_datetime = env.working_datetime
    working_datetime_date  = working_datetime.date()
if args.today:
    working_datetime_date = datetime.today()
    working_date_str      = working_datetime_date.strftime('%Y-%m-%d')
    working_datetime      = datetime.combine(working_datetime_date, datetime.min.time())
if args.date:
    working_datetime      = datetime.strptime(args.date, "%Y-%m-%d")
    working_datetime_date = working_datetime.date()
    working_date_str      = working_datetime_date.strftime('%Y-%m-%d')

if args.no_elec:
    elec_usage = None
else:
    try:
        elec_usage = electron.get_elecUsage(working_datetime_date)
        if not elec_usage:
            logger.warning('电表数据获取错误，请检查')
    except Exception as e:
        logger.error(f'{e}')

logger.info(f'working date is: {working_datetime_date}')

dir_str     = f"{env.proj_dir}/et/{working_datetime_date.strftime('%m%d')}日报表"
source_file = env.source_file
save2file   = f"{dir_str}/2025年日报表.xlsx"

mt = get_meituanSum(working_datetime)
if not mt:
    logger.warning('美团数据没拿到')
dy = get_douyinSum(working_datetime)
if not dy:
    logger.warning('抖音数据没有拿到')
english = resolve_operation_data(working_datetime)
if not english:
    logger.warning('运营数据没拿到')

ota_update  = ThirdParty.ota_update
repeat_ids = ThirdParty.check_unique(working_date_str)

d_status = 0
for thirdtype, ids in repeat_ids.items():
    for id in ids:
        logger.info(f'正在删除{thirdtype}:{id}')
        delete_status = ThirdParty.delete(id)
        if delete_status == 1:
            logger.warning(f'删除失败，不添加')
            d_status = 1
if d_status == 0:
    logger.info('删除成功')
    ota_update(ota_name='DOUYIN', date_obj=working_datetime, income=dy)
    ota_update(ota_name='MEITUAN', date_obj=working_datetime, income=mt)

specialFee_list, special_sum = specialFee.get_specialFee(working_datetime)        #可以排序

working_sheetname = f'{working_datetime.month}月'
wb = load_workbook(source_file)

if working_sheetname in wb.sheetnames:
    ws = wb[working_sheetname]
    wb.active = ws
    logger.info(f'操作的worksheet{wb}:{ws}')
else:
    logger.error(f'当前月份的工作表不存在，请手动检查.{working_sheetname}')

def load_data():
    chinese = {}
    data_pure = {
        "用电量" : elec_usage ,
        "美团"   : mt   ,
        "抖音"   : dy
    }

    cn_en_map = {
        "网费充值"  :     "amountFee",
        "提现本金"  :     "withdrawPrincipal",
        "找零"      :     "checkoutDeposit",
        "零售"      :     "retail",
        "水吧"      :     "waterBar",
        "代购"      :     "agent",
        "退款"      :     "totalRefundFee",
        "报销"      :     None,
        "在线支付"  :     "onlineIn",
        "奖励金"    :     "awardFee",
        "卡券"      :     "cardVolumeFee",
        "特免"      :     "specialFree",
        "网费消耗"  :     "totalConsumeNetworkFee",
        "上机人次"  :     "onlineTimes",
        "上机时长"  :     "duration",
        "点单率"    :     "orderRate",
        "新会员"    :     "newMember"
    }

    for cn_name, eng_name in cn_en_map.items():
        if eng_name and eng_name in english:
            data_pure[cn_name] = english[eng_name]
        else:
            data_pure[cn_name] = 0

    data_pure["退款"]       = -data_pure["退款"] 
    if data_pure['退款'] == 0 :
        data_pure['退款'] = None

    data_pure["上机时长"]   = round(data_pure["上机时长"] /60, 2)
    return data_pure

def insert_data(data_pure):
    target_row = electron.get_row_by_date(ws, working_datetime_date)
    
    logger.info(f'target_row is {target_row}')
    import json
    fmt_json = json.dumps(data_pure, ensure_ascii=False, indent=4)
    logger.info(f'\n最终要写入et的数据:\n{fmt_json}')

    for col in ws.iter_cols(min_row=2, max_row=2, min_col=1, max_col=29):
        header = col[0].value  # 第二行的列标题
        if header in data_pure:
            ws.cell(row=target_row, column=col[0].column, value=data_pure[header])

    date_str = working_datetime.strftime("%-m月%-d日")
    #%-m 和 %-d 中的减号用于去除月份和日期中的前导零。请注意，这种用法在某些操作系统（如 Unix/Linux）上有效，但在 Windows 上可能不被支持。

    #ws["B36"].value = f"芜湖张家山店{date_str}营业状况"
    ws["C36"].value = f"芜湖张家山店{date_str}营业状况"

font_wenquan = Font(name='WenQuanYi Zen Hei',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000'
        )

font_yahei = Font(name='Microsoft YaHei',
                size= 8,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000'
        )
ws['G37'].font = font_yahei

yahei_inline = InlineFont(rFont='Microsoft YaHei',
                sz=11,
                color='FF000000'
        )
yahei_22 = InlineFont(rFont='Microsoft YaHei',
                sz=22,
                color='FFFF0000'
                    )
yahei_red_8 = InlineFont(rFont='Microsoft YaHei',
                sz=8,
                color='FFFF0000'
                    )
b = TextBlock

def old_special_mark(special_data):
    left_alignment = Alignment(horizontal='left')
    center_alignment = Alignment(horizontal='center')

    thin    = Side(border_style="thin", color="888888")
    medium  = Side(border_style="medium", color="000000")
    dash    = Side(border_style="dashed", color="000000")
    dotted  = Side(border_style="dotted", color="000000")

    merged_ranges = ws.merged_cells.ranges

    for row in range(37, 57):  # 37 到 57 行

        merge_range_GI = f'H{row}:K{row}'
        merge_range_GH = f'H{row}:J{row}'

        # 检查 G{row}:I{row} 是否已被合并
        if any(str(merged_range) == merge_range_GI for merged_range in merged_ranges):
            ws.unmerge_cells(merge_range_GI)
            ws.merge_cells(merge_range_GH)
        left  = ws[f'H{row}']
        right = ws[f'I{row}']
        weft  = ws[f'K{row}']
        weft.value  = None
        #breakpoint()
        #right.value = None
        weft.border  = Border(left = thin, bottom=dotted )
        left.border  = Border(bottom=dotted)
        right.border = Border(left=None, right=medium, bottom=dotted)
        right.font  = copy(ws['H37'].font)
        weft.font   = copy(ws['H37'].font)
 
    bottom = 57
    ws[f'H{bottom}'].border = Border(left=thin,bottom=medium)
    ws[f'I{bottom}'].border = Border(bottom=medium)
    ws[f'J{bottom}'].border = Border(bottom=medium,right=medium)
    ws[f'K{bottom}'].border = Border(bottom=medium,right=medium)

    for index, item in enumerate(special_data):
        row = 37 + index
        if row > 56:
            break  # 超过 G57，停止插入
        reason, value  = item.split('：')

        #rich_string = CellRichText(
        #    b(yahei_inline, f'{reason}'),
        #    b(yahei_red_8, f'{value}'),
        #    b(yahei_red_8, 'sss')
        #    )
        cell = ws.cell(row=row, column=7)
        cell.value = item
        #cell = rich_string
        cell.alignment = left_alignment

        yuan = ws.cell(row=row, column=9)
        yuan.value = "元"
        yuan.alignment = center_alignment


def special_mark(ws, special_data, start_col, end_col, start_row=37, end_row=57):
    left_alignment = Alignment(horizontal='left')
    center_alignment = Alignment(horizontal='center')

    thin    = Side(border_style="thin", color="888888")
    medium  = Side(border_style="medium", color="000000")
    dotted  = Side(border_style="dotted", color="000000")

    merged_ranges = ws.merged_cells.ranges

    start_col_index = column_index_from_string(start_col)
    end_col_index   = column_index_from_string(end_col)
    merge_right_col_index = end_col_index - 1  # 合并区域去掉“元”列

    merge_left_col  = start_col
    merge_right_col = get_column_letter(merge_right_col_index)
    yuan_col        = end_col  # 最右边是“元”列
    for row in range(start_row, end_row):
        for col_letter in [start_col ]:  # 例如 H 和 K
            cell = ws[f'{col_letter}{row}']
            cell.value = None

    for row in range(start_row, end_row):
        merge_range = f'{merge_left_col}{row}:{merge_right_col}{row}'
        full_range  = f'{merge_left_col}{row}:{yuan_col}{row}'

        # 如果整行已合并，拆分后再合并前几列
        if any(str(rng) == full_range for rng in merged_ranges):
            ws.unmerge_cells(full_range)
        ws.merge_cells(merge_range)

        left_cell  = ws[f'{merge_left_col}{row}']
        yuan_cell  = ws[f'{yuan_col}{row}']
        mid_cell   = ws.cell(row=row, column=merge_right_col_index)

        yuan_cell.value = None

        left_cell.border = Border(left = thin, bottom=dotted)
        mid_cell.border  = Border(left=None, right=None, bottom=dotted)
        yuan_cell.border = Border(bottom=dotted,right=medium)

        base_font = copy(ws[f'{merge_left_col}{start_row}'].font)
        left_cell.font   = base_font
        mid_cell.font    = base_font
        yuan_cell.font   = base_font

    # 设置最后一行底部边框
    bottom = end_row - 1
    ws[f'{merge_left_col}{bottom}'].border  = Border(left=thin, bottom=dotted)
    ws[f'{merge_right_col}{bottom}'].border = Border(bottom=dotted)
    ws[f'{yuan_col}{bottom}'].border        = Border(bottom=dotted, right=medium)

    # 写入数据
    for index, item in enumerate(special_data):
        row = start_row + index
        if row >= end_row:
            break
        if '：' in item:
            reason, value = item.split('：', 1)
        else:
            reason, value = item, ''

        text_cell = ws.cell(row=row, column=start_col_index)
        text_cell.value = item
        text_cell.alignment = left_alignment

        yuan_cell = ws.cell(row=row, column=end_col_index)
        yuan_cell.value = "元"
        yuan_cell.alignment = center_alignment


def A1Bug():
    '''
        A1的标题样式有bug，单独执行这个函数来处理
    '''
    red_font = Font(color="FF0000")
    rich_text = CellRichText(
        [
            "网费收入",
            TextBlock(InlineFont(red_font), "(网费+提现+找零)"),  # 使用 InlineFont 包装
        ]
    )
    ws['A1'] = rich_text
    font1 = Font(bold=True, size=16)
    ws['A1'].font = font1

def handle_headers(ws):
    mcells = [ "f1", "j1", "n1", "s1", "w1", "ac1" ]

    for cell in mcells:
        text = ws[cell].value
        start_index = text.find("(")
        end_index = text.find(")")

        if start_index != -1 and end_index != -1:
            # 创建 Font 对象，设置字体大小
            small_font = Font(size=8)
            # 创建 CellRichText 对象
            rich_text = CellRichText(
                [
                    text[:start_index],  # 括号前的文本
                    TextBlock(InlineFont(small_font), text[start_index:end_index + 1]),  # 括号内的文本，小字
                    text[end_index + 1:],  # 括号后的文本
                ]
            )
        else:
            # 如果没有括号，直接设置文本
            rich_text = CellRichText([text])
        ws[cell] = rich_text
        A1Bug()

def save(path):
    try:
        os.makedirs(dir_str, exist_ok=True)
        print(f"目录 '{dir_str}' 创建成功")
    except Exception as e:
        print(f"创建目录时发生错误: {e}")
    try:
        shutil.move(source_file, f"{source_file}.old")
        wb.save(source_file)
        wb.save(path)
    except FileNotFoundError:
        print(f"文件 {path} 或目录 {os.path.dirname(source_file)} 不存在")
    except Exception as e:
        print(f"复制文件 {path} 出错：{e}")

#判断specialFeesum和specialFee是否一致I#
data_pure = load_data()
if data_pure['特免'] != special_sum:
    logger.warn(f"特免金额不匹配，请检查!!运营数据中是{data_pure['特免']},订单列表中计算出来是{special_sum}")

insert_data(data_pure)


special_mark(ws = ws, special_data = specialFee_list, start_col='H', end_col='K' )
handle_headers(ws)
save(save2file)
