import shutil
import os
import argparse
from datetime   import datetime, date, timedelta
from copy       import copy

from openpyxl                import load_workbook
from openpyxl.styles         import Font, Border, Side,  Alignment
from openpyxl.cell.text      import InlineFont
from openpyxl.cell.rich_text import CellRichText, TextBlock, TextBlock

from tools              import env
from openpyxl.utils import range_boundaries
from tools import logger

logger = logger.getlogger(__name__)

def next_tuesday_or_saturday(last_datetime):
    """
    获取下一个最近的周二或周六的日期。

    参数：
        last_datetime (datetime): 输入的日期对象。

    返回：
        datetime: 下一个周二或周六的日期。
    """
    # 定义周二和周六的星期索引
    TUESDAY = 1
    SATURDAY = 5

    # 获取输入日期的星期索引
    current_weekday = last_datetime.weekday()
    print('current_weekday', current_weekday)
    # 计算距离下一个周二和周六的天数
    days_until_tuesday = (TUESDAY - current_weekday + 7) % 7
    days_until_saturday = (SATURDAY - current_weekday + 7) % 7

    # 如果计算结果为0，说明输入日期就是目标日之一，需要跳过当前日期
    if days_until_tuesday == 0:
        days_until_tuesday = 7
    if days_until_saturday == 0:
        days_until_saturday = 7

    # 返回距离最近的目标日期
    if days_until_tuesday < days_until_saturday:
        return last_datetime + timedelta(days=days_until_tuesday)
    else:
        return last_datetime + timedelta(days=days_until_saturday)

def find_last_filled_cell(worksheet, cell_range):
    """
    在指定的单元格范围内，找到最后一个有数值的单元格。

    参数：
        worksheet: openpyxl 的工作表对象。
        cell_range: 字符串，表示单元格范围，例如 'A1:C10'。

    返回：
        最后一个有数值的单元格对象；如果范围内没有数值，则返回 None。
    """
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    last_filled_cell = None

    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row,
                                   min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                last_filled_cell = cell

    return last_filled_cell

def append_date(ws, start_row, date_value):
    """
    在指定的单元格范围内插入日期。

    参数：
    ws : openpyxl.worksheet.worksheet.Worksheet
        要操作的工作表对象。
    cell_range : str
        目标单元格范围，例如 'B2:D4'。
    date_value : datetime
        要插入的日期值，datetime 对象。
    """
            #cell.number_format = 'yyyy-mm-dd'  # 设置日期格式

    for i in range(4):
        worksheet[f'A{start_row+i+1}'].value = date_value


market_et_path = f'{env.proj_dir}/et/0325市调表.xlsx'
save_path = f'{env.proj_dir}/et/market_et_test.xlsx'
workbook = load_workbook(market_et_path)
worksheet = workbook.active
target = find_last_filled_cell(worksheet, 'A1:A50')
target_row = target.row

workday = next_tuesday_or_saturday(target.value)

#if date.today().weekday() == 1 or date.today().weekday() == 5:
#    workday = date.today()

append_date(worksheet, target.row, workday)

print(worksheet['A34'].value)

def copy_firm(column_letter, start_row):
    '''
    column_letter   : string,
    start_row       : int
        '''

    for i in range(4):
        worksheet[f'{column_letter}{start_row+4}'].value = worksheet[f'{column_letter}{start_row}'].value
        start_row -= 1

copy_firm('B', target_row)
copy_firm('D', target_row)
copy_firm('F', target_row)

workbook.save(save_path)
