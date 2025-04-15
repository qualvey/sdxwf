
import json
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from datetime import datetime, timedelta, time
from urllib.parse import urlencode,urlunparse

from worktime import get_userid, consume_duration
from tools import env, iheader, find, logger

logger = logger.get_logger(__name__)

userid_map = get_userid.userids
#dict

work_time_xl = f"{env.proj_dir}/et/2025年3月工时表.xlsx"
workbook = load_workbook(work_time_xl)
worksheet = workbook.active

#today = datetime.today()-timedelta(days=8)
today = datetime.today()
today_datetime = datetime.combine(today, time.min)
breakpoint()

duration_sum = {}

def get_previous_week_range(dt):
    # 计算当前日期是星期几，星期一为0，星期日为6
    current_weekday = dt.weekday()
    # 计算上周日距离当前日期的天数
    days_since_last_sunday = current_weekday + 1
    # 计算上周日的日期
    last_sunday = dt - timedelta(days=days_since_last_sunday)
    # 计算上周一的日期
    last_monday = last_sunday - timedelta(days=6)
    last_monday = datetime.combine(last_monday, time.min)
    week_dates = [last_monday + timedelta(days=i) for i in range(7)]
    return week_dates

weekdays = get_previous_week_range(today)

obj_mon = weekdays[0]
obj_sun = datetime.combine(weekdays[6], time.max)
mon = str(obj_mon)
sun = str(obj_sun)

logger.debug(f'请求的日期范围:{str(mon),str(sun)}' )
print(f'请求的日期范围:{str(mon),str(sun)}' )


def get_total_wduration(data_json):
    get_total_wduration.call_count+=1
    sum = 0
    if data_json['data']['total'] == 7:
        data_list = data_json['data']['rows'] 
    else:
        print("not 7days") 
        data_list = data_json['data']['rows'] 

    for item in data_list:
        name  = item['userName']
        print(name)
        sum += item['duration']
    return sum

get_total_wduration.call_count = 0

def  write_to_file(response_data):
    try:
        new_data = response_data.json()
        duration = get_total_wduration(new_data)
        file_path="./worktime.json"
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                try:
                    data = json.load(file)
                except json.JSONDecodeError:
                    # 如果文件为空或内容不是有效的 JSON，初始化为空列表
                        data = []
        except FileNotFoundError:
            data = []
        if isinstance(data, list):
            # 将新数据追加到列表中
            data.append(new_data)
        else:
            # 如果数据不是列表，可能是字典或其他结构，根据需要处理
            print("现有数据不是列表，无法追加新数据。")
            # 视情况决定是将数据转换为列表，还是采取其他措施
            # 例如，将现有数据和新数据放入一个新的列表中
            data = [data, new_data]
        with open(file_path, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=4)
        return new_data
    except requests.exceptions.JSONDecodeError:
        print("响应内容不是有效的 JSON 格式")


scheme = 'https'
netloc = 'hub.sdxnetcafe.com'
path = '/api/admin/work/attendance/list'
params =''
fragment = ''
def request_get(userid):
    query = {
        "startTm":  mon,
        "endTm":    sun,
        "page": 1,
        "limit": 10,
        "userId": userid,
        "branchId": "a92fd8a33b7811ea87766c92bf5c82be"
    }

    query_string = urlencode(query)
    url = urlunparse((scheme, netloc, path, params, query_string, fragment))
    response_user  = requests.get(url, headers=iheader.headers)
    return response_user.json()

def work_duration_all():
    work_duration_json_map = {}
    work_duration_time_map = {}

    for name, id in userid_map.items():
        work_duration_json_map[name] = request_get(id)
        #此变量有一个用户每天的数据
        duration_aday = get_total_wduration(work_duration_json_map[name])
        print(f'调用了{get_total_wduration.call_count}次')
        duration_sum[name] = duration_aday

def get_json_person():
    members_json = {}
    for name, id in userid_map.items():
        data_json =  request_get(id)
        members_json[name] =  data_json
    return members_json

yahei_red_8 = InlineFont(rFont='Microsoft YaHei',
                color='FFFF0000'
                         )


b = TextBlock
def to_xl(worksheet, json):

    worktime_ist = json['data']['rows']
    name         = json['data']['rows'][0]['userName']
    id_row = None
    cols = list(worksheet.iter_cols(min_row=1, max_row=20, min_col=1, max_col=1))
    for cell in cols[0]:
        cell_value = str(cell.value)
        if name in cell_value:
            id_row = cell.row
            break
    if id_row == None:
        return 1
    date_format = "%Y-%m-%d"

    for i in worktime_ist:

        date_object = datetime.strptime(i['dutyDate'], date_format)
        datecell = find.get_cell_by_datetime(worksheet, date_object, start_cell="B2", end_cell="AG2")
        column_index = datecell.column
        column_letter = get_column_letter(column_index)
        target_cell = worksheet[f'{column_letter}{id_row}']
        
        if i['duration'] == 0:
            rich_string = CellRichText(b(yahei_red_8, f'休'))
            target_cell.value = rich_string
        else:
            target_cell.value = i['duration']

work_duration_all()

members_json = get_json_person()

with open(f'{env.proj_dir}/worktime/members_json.json', 'w') as cache:
    json.dump(members_json, cache, ensure_ascii=False, indent=4)

for name, user_json in  members_json.items():

    if len(user_json['data']['rows']) > 0:
       to_xl(worksheet, user_json)

workbook.save(work_time_xl)

sum = 0
for  num in duration_sum.values():
    sum += num

if __name__ == "__main__":
    get_value_days = consume_duration.get_value_days
    consume_duration = get_value_days(weekdays)
    print("上机时长",consume_duration)
    print("总工时",sum)
    print('人效', consume_duration/sum)
