from openpyxl import load_workbook
from tools import env

proj_dir    = env.project_dir
target_file = f"{project_dir}/2025年日报表.xlsx"
exclude_columns = ["E", "I", "M", "R", "V"]

wb = load_workbook(target_file)
ws = wb.active

def delete_cells(ws, start_cell, end_cell):
    """
    有效范围: B2, AB33 (这里修改了，以匹配实际调用)
    """
    for row in ws.iter_rows(min_row=ws[start_cell].row, max_row=ws[end_cell].row,
                           min_col=ws[start_cell].column, max_col=ws[end_cell].column):
        for cell in row:
            if cell.column_letter not in exclude_columns:
                cell.value = None

delete_cells(ws, "B13", "AB13")
wb.save(target_file)
