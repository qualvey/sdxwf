from openpyxl import load_workbook
import os
import tkinter as tk
from datetime import datetime, date, timedelta
import json
from tools import env
from tools.logger import get_logger

import sqlite3

logger = get_logger(__name__)

def init_db():
# 连接到 SQLite 数据库（如果文件不存在，会自动创建）
    conn = sqlite3.connect(f'{env.proj_dir}/energy_data.db')
    cursor = conn.cursor()
# 创建表（如果不存在）
    create_table_query = '''
    CREATE TABLE IF NOT EXISTS daily_energy_consumption (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_date DATE NOT NULL UNIQUE,
        energy_consumed REAL NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    '''
    cursor.execute(create_table_query)
    conn.commit()

def is_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def update_sql_elec(record_date, energy_consumed):
    init_db()

    conn = sqlite3.connect(f'{env.proj_dir}/energy_data.db')
    cursor = conn.cursor()
    '''
        record_date <datetime.datetime>
        energy_consumed <float>
    '''
    record_date = record_date.strftime('%Y-%m-%d')
    try:
        insert_query = '''
        INSERT INTO daily_energy_consumption (record_date, energy_consumed)
        VALUES (?, ?)
        ON CONFLICT(record_date) DO UPDATE SET energy_consumed = excluded.energy_consumed;
        '''
        cursor.execute(insert_query, (record_date, energy_consumed))
        conn.commit()
        print("数据插入成功")
    except sqlite3.Error as err:
        print(f"插入数据时发生错误: {err}")

# 定义查询日期
def query_data(date_str):

    conn = sqlite3.connect(f'{env.proj_dir}/energy_data.db')
    cursor = conn.cursor()

    cursor.execute('SELECT energy_consumed FROM daily_energy_consumption WHERE record_date = ?', (date_str,))

    rows = cursor.fetchall()
    for row in rows:
        print(row)
        return(row[0])

elecSheet   = env.elecUsage_file

wb = load_workbook(elecSheet)
#data_only = True 获取数据而不是公式
#dataonly不能保存，保存就破坏公式了
ws = wb.active

def get_row_by_date(worksheet,search_value,start_cell="A1",end_cell="A35"):
    """
    参数：工作表，查找值<datatime>, 开始和终止的cell位置
    返回：第一个匹配到的cell行号
    """
    try:
        min_row, max_row = worksheet[start_cell].row, worksheet[end_cell].row
        min_col, max_col = worksheet[start_cell].column, worksheet[end_cell].column

        for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if isinstance(cell.value, datetime):  # 确保是 datetime 类型
                    if cell.value.date() == search_value:  # 对比日期部分
                        print('命中')
                        print(cell.row)
                        return cell.row  # 直接返回行号
        return None  # 未找到匹配值

    except KeyError:
        print(f"错误: {start_cell} 或 {end_cell} 可能超出了工作表范围")
        return None

class UserCancledException(Exception):
    pass
def write_elecxl(elec_usage, target_row, destination):
    if not target_row :
        print('电表有问题，未获取正确的位置')
        return 
    ws[f"B{target_row}"].value = elec_usage
    wb.save(elecSheet)
    wb.save(destination)

def get_elecUsage(datetime_obj):
    #这个可以用GUI
    #ele_usage   = env.configjson['elecUsage']
    #函数，获取输入的值并赋给变量
    def on_enter_pressed(a):
        user_input = entry.get()  # 获取用户输入
        print(f"输入了用电量: {user_input}")  # 输出变量值
        # 如果需要将输入值存储为变量，可以在这里进一步处理
        ele_usage.set(user_input)  # 更新 my_variable 的值
        root.quit()
    search_date = datetime_obj
    if_exitst = query_data(search_date.strftime('%Y-%m-%d'))
    if if_exitst:
        logger.warning('该日期已经存在电表数据，如果要修改，请手动操作sqlite')
        ele_usage = if_exitst
    else:

        # 创建主窗口
        root = tk.Tk()
        root.title("输入电表数据")
        root.geometry('300x300')

        # 创建输入框
        entry = tk.Entry(root)
        entry.pack(pady=10)

# 创建按钮，点击时获取输入框中的值
        button = tk.Button(root, text=f"请输入用电量{search_date}", command=on_enter_pressed)
        button.pack(pady=10)

# 创建一个显示变量的标签
        ele_usage = tk.StringVar()
        label = tk.Label(root, textvariable=ele_usage)
        label.pack(pady=10)
        entry.bind('<Return>', on_enter_pressed)

# 运行主循环
        root.mainloop()
        ele_usage = float(ele_usage.get())
        update_sql_elec(datetime_obj, ele_usage)
    previous_day = datetime_obj-timedelta(days=1)
    previous_day_str = previous_day.strftime('%Y-%m-%d')
    previous_day_value = query_data(previous_day_str)
    logger.info(previous_day_value)
    if previous_day_value:
        result = ele_usage - previous_day_value
    else:
        logger.warn(f'前一天<{previous_day.__str__()}>的数据不存在,是否添加? y<int>/n')
        user_input = input(f'输入y紧跟电表数据\n').strip().lower()
        if user_input.startswith("y") and is_float(user_input[1:]):
            pre_day_value = float(user_input[1:])
            result = ele_usage - pre_day_value
            conn = sqlite3.connect(f'{env.proj_dir}/energy_data.db')
            cursor = conn.cursor()
            record_date = previous_day.strftime('%Y-%m-%d')
            try:
                insert_query = '''
                INSERT INTO daily_energy_consumption (record_date, energy_consumed)
                VALUES (?, ?)
                ON CONFLICT(record_date) DO UPDATE SET energy_consumed = excluded.energy_consumed;
                '''
                cursor.execute(insert_query, (record_date, pre_day_value))
                conn.commit()
                print("数据插入成功")
            except sqlite3.Error as err:
                print(f"插入数据时发生错误: {err}")
        else: 
            logger.warn(f'user_input.startswith(\'y\'):{user_input.startswith('y')}, user_input[1:].isdigit():{user_input[1:]}')
            raise UserCancledException('用户取消，电表模块退出')


    result = result*80
    return int(result) 

    cursor.close()

    conn.close()
